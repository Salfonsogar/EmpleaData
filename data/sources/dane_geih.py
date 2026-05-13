import openpyxl
import re
from typing import Dict, List, Tuple, Optional

from core.constants import AÑOS

# Mapping DANE city names → project city names
CIUDAD_MAP = {
    "Bogotá D.C.": "Bogotá",
    "Medellín A.M.": "Medellín",
    "Cali A.M.": "Cali",
    "Barranquilla A.M.": "Barranquilla",
    "Cartagena": "Cartagena",
    "Bucaramanga A.M.": "Bucaramanga",
    "Cúcuta A.M.": "Cúcuta",
    "Pasto": "Pasto",
    "Sincelejo": "Sincelejo",
    "Santa Marta": "Santa Marta",
    "Valledupar": "Valledupar",
    "Montería": "Montería",
    "Riohacha": "Riohacha",
    "Quibdó": "Quibdó",
    "Arauca": "Arauca",
    "Leticia": "Leticia",
}

# Cities with their region for fallback imputation
CIUDAD_REGION = {
    "Barranquilla": "Caribe", "Cartagena": "Caribe", "Santa Marta": "Caribe",
    "Valledupar": "Caribe", "Montería": "Caribe", "Sincelejo": "Caribe",
    "Riohacha": "Caribe", "Bogotá": "Triángulo de Oro", "Medellín": "Triángulo de Oro",
    "Cali": "Triángulo de Oro", "Bucaramanga": "Santanderes", "Cúcuta": "Santanderes",
    "Quibdó": "Fronterizo", "Arauca": "Fronterizo", "Leticia": "Fronterizo",
    "Pasto": "Fronterizo",
}

# DANE regions for Amazon/Orinoquía imputation
FRONTERIZA_CITIES = {"Arauca", "Leticia", "Quibdó", "Pasto"}


def _find_sheet(wb) -> str:
    candidates = [s for s in wb.sheetnames if "Año_móvil" in s or "año_móvil" in s]
    if candidates:
        return candidates[0]
    candidates = [s for s in wb.sheetnames if "areas trim movil" in s.lower()]
    if candidates:
        return candidates[0]
    # NEW FORMAT: "23 ciudades" or "13 ciudades A.M."
    for name in ["23 ciudades", "13 ciudades A.M.", "13 ciudades"]:
        if name in wb.sheetnames:
            return name
    raise ValueError("No se encontró una hoja con datos por ciudad")


def _find_calendar_year_columns(ws) -> Dict[int, int]:
    headers = {}
    latest_2026_col = None
    latest_2026_order = -1

    # NEW FORMAT: Check if row 2 has "Ene - Dic YYYY" pattern (columns B-G)
    row2 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    if row2:
        for col_idx, val in enumerate(row2[0], start=1):
            if val and isinstance(val, str):
                m = re.match(r'Ene\s*-\s*Dic\s+(\d{2,4})', val.strip())
                if m:
                    year = int(m.group(1))
                    if year < 100:
                        year += 2000
                    if year in AÑOS:
                        headers[year] = col_idx

    # OLD FORMAT: search through multiple rows
    if not headers:
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
            for c in row:
                if c.value and isinstance(c.value, str):
                    v = c.value.strip()
                    m = re.match(r'Ene\s*-\s*Dic\s+(\d{2,4})', v)
                    if m:
                        year = int(m.group(1))
                        if year < 100:
                            year += 2000
                        if year in AÑOS:
                            headers[year] = c.column

                    if v and v.endswith("26") and any(mes in v for mes in ["Ene", "Feb", "Mar", "Abr"]):
                        months = {"Ene": 1, "Feb": 2, "Mar": 3, "Abr": 4, "May": 5, "Jun": 6,
                                  "Jul": 7, "Ago": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dic": 12}
                        for prefix, order in months.items():
                            if v.startswith(prefix):
                                if order > latest_2026_order:
                                    latest_2026_order = order
                                    latest_2026_col = c.column
                                break

        if 2026 not in headers and latest_2026_col is not None:
            headers[2026] = latest_2026_col

    return headers


def _find_to_row(ws, start_row: int, max_look: int = 8) -> Optional[int]:
    for r in range(start_row, start_row + max_look):
        val = ws.cell(row=r, column=1).value
        if val and isinstance(val, str):
            v = val.strip().lower()
            if v in ('to',) or v.startswith('tasa de ocupación'):
                return r
    return None


def _find_city_blocks(ws) -> List[Tuple[str, int]]:
    blocks = []
    # NEW FORMAT: Data starts at row 5 (city in column A, values in B-G)
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row or 1000, min_col=1, max_col=1, values_only=False):
        c = row[0]
        if c.value and isinstance(c.value, str):
            name = c.value.strip()
            if name in CIUDAD_MAP:
                blocks.append((name, c.row))
    # OLD FORMAT fallback: search from row 1
    if not blocks:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1000, min_col=1, max_col=1, values_only=False):
            c = row[0]
            if c.value and isinstance(c.value, str):
                name = c.value.strip()
                if name in CIUDAD_MAP:
                    blocks.append((name, c.row))
    return blocks


def extract_tasa_ocupacion(filepath: str) -> Dict[str, List[float]]:
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheet_name = _find_sheet(wb)
    ws = wb[sheet_name]

    year_cols = _find_calendar_year_columns(ws)
    if not year_cols:
        wb.close()
        raise ValueError(f"No se encontraron columnas de año calendario en '{sheet_name}'")

    blocks = _find_city_blocks(ws)
    if not blocks:
        wb.close()
        raise ValueError(f"No se encontraron ciudades conocidas en '{sheet_name}'")

    # NEW FORMAT: values are in same row as city name (columns B-G = 2-7)
    # Check if first block has data in same row
    result = {}
    first_city, first_row = blocks[0]
    first_row_values = [ws.cell(row=first_row, column=c).value for c in range(2, 8)]
    new_format = any(v is not None for v in first_row_values)

    for dane_name, city_row in blocks:
        proj_name = CIUDAD_MAP[dane_name]

        if new_format:
            # NEW FORMAT: values in same row
            values = []
            for year in AÑOS:
                col = year_cols.get(year)
                if col is None:
                    values.append(None)
                else:
                    cell_val = ws.cell(row=city_row, column=col).value
                    if cell_val is not None:
                        try:
                            values.append(round(float(cell_val), 1))
                        except (ValueError, TypeError):
                            values.append(None)
                    else:
                        values.append(None)
        else:
            # OLD FORMAT: values in TO row below city
            to_row = _find_to_row(ws, city_row + 1)
            if to_row is None:
                continue

            values = []
            for year in AÑOS:
                col = year_cols.get(year)
                if col is None:
                    values.append(None)
                else:
                    cell_val = ws.cell(row=to_row, column=col).value
                    if cell_val is not None:
                        try:
                            values.append(round(float(cell_val), 1))
                        except (ValueError, TypeError):
                            values.append(None)
                    else:
                        values.append(None)
        result[proj_name] = values

    wb.close()
    return result


def impute_regional_mean(data: Dict[str, List[float]]) -> Dict[str, List[float]]:
    result = dict(data)

    fronterizo_vals = [[] for _ in AÑOS]
    for city, vals in result.items():
        if city in FRONTERIZA_CITIES:
            for i, v in enumerate(vals):
                if v is not None:
                    fronterizo_vals[i].append(v)

    for city in ["Arauca", "Leticia"]:
        if city not in result or all(v is None for v in result.get(city, [])):
            if city not in result:
                result[city] = []
            imputed = []
            for i in range(len(AÑOS)):
                if fronterizo_vals[i]:
                    imputed.append(round(sum(fronterizo_vals[i]) / len(fronterizo_vals[i]), 1))
                else:
                    imputed.append(None)
            result[city] = imputed

    return result
